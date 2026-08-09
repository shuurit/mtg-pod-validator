"""
One-time structural migration: adds a "Playgroup Game ID" column (Y) to the
current season's Game Log tab, right after "Game Calculated Deck Strength"
(X). Header only -- existing data rows are left untouched; see
backfill_game_ids.py for populating it on already-logged rows, and
add_game.py (now accepts an optional playgroupGameId) for new ones.

Column Y is a deliberate exact-match reference key: computeLoggedMatches in
app.js can look a playgroup.gg game up by ID directly instead of inferring
"same game" from player names + date + commander, which is what actually
caused the Game 13 bug fixed earlier this session. Only ever populated for
games playgroup.gg itself reports (backfill_game_ids.py only matches
against the live, current-active-league /playgroup-games response) -- a
manually-entered or historical non-league game simply has this column
blank forever, and still falls back to the existing heuristic matching.

Usage:
    python scripts/add_game_id_column.py
"""
import openpyxl

from xlsx_recalc import recalc
from season_config import CURRENT_SEASON_SHEET, XLSX_PATH

HEADER_ROW = 2
NEW_COL = "Y"
CATEGORY_LABEL = "Playgroup Game ID"
COLUMN_LABEL = "Playgroup Game ID"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb[CURRENT_SEASON_SHEET]

    existing = ws[f"{NEW_COL}{HEADER_ROW}"].value
    if existing:
        raise RuntimeError(f"{NEW_COL}{HEADER_ROW} already has a value ({existing!r}) -- column already added?")

    import copy
    row1_style_src = ws["V1"]
    row2_style_src = ws["V2"]

    cell1 = ws[f"{NEW_COL}1"]
    cell1.value = CATEGORY_LABEL
    if row1_style_src.has_style:
        cell1._style = copy.copy(row1_style_src._style)

    cell2 = ws[f"{NEW_COL}{HEADER_ROW}"]
    cell2.value = COLUMN_LABEL
    if row2_style_src.has_style:
        cell2._style = copy.copy(row2_style_src._style)

    wb.save(XLSX_PATH)
    print(f'Added "{COLUMN_LABEL}" header at {NEW_COL}1/{NEW_COL}{HEADER_ROW} in {CURRENT_SEASON_SHEET}.')

    recalc(XLSX_PATH)


if __name__ == "__main__":
    main()
