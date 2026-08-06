"""
Shared LibreOffice-headless recalculation for deck-strength.xlsx.

Used both by add_game.py (after writing a new game's rows) and by
recalculate.py (a standalone, manually-triggerable pass -- useful after any
script that edits the workbook via direct XML surgery instead of through
openpyxl, since that path never gets a real Excel/LibreOffice recalc pass
on its own).
"""
import os
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      Dim oSheets As Object
      Dim oSheet As Object
      Dim oPivots As Object
      Dim i As Integer, j As Integer

      oSheets = ThisComponent.getSheets()
      For i = 0 To oSheets.getCount() - 1
        oSheet = oSheets.getByIndex(i)
        oPivots = oSheet.getDataPilotTables()
        For j = 0 To oPivots.getCount() - 1
          oPivots.getByIndex(j).refresh()
        Next j
      Next i

      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def recalc(xlsx_path):
    # Plain `soffice --convert-to` does NOT reliably force a full formula
    # recalculation -- confirmed the hard way: it silently produced a file
    # where every formula cell across the whole workbook was simply missing
    # (not an error value, just absent), because openpyxl's save() had
    # stripped their cached values and convert-to never recomputed them.
    # A macro that explicitly calls calculateAll() before store() is the
    # reliable way to force it. Pivot tables (DataPilotTables in LibreOffice's
    # API) are a separate cache calculateAll() doesn't touch on its own --
    # refreshed explicitly here so Deck Win Rates doesn't go stale again.
    abs_path = os.path.abspath(xlsx_path)

    with tempfile.TemporaryDirectory(prefix="lo_profile_") as profile_dir:
        profile_url = Path(profile_dir).as_uri()

        boot = subprocess.run(
            ["soffice", "--headless", "--terminate_after_init", f"-env:UserInstallation={profile_url}"],
            capture_output=True, text=True, timeout=60,
        )
        if boot.returncode != 0:
            print(boot.stderr, file=sys.stderr)
            raise RuntimeError("LibreOffice failed to initialize its profile")

        macro_dir = Path(profile_dir) / "user" / "basic" / "Standard"
        if not macro_dir.exists():
            raise RuntimeError("LibreOffice did not create a usable profile directory")
        (macro_dir / "Module1.xba").write_text(RECALCULATE_MACRO)

        result = subprocess.run(
            [
                "soffice", "--headless", "--norestore",
                f"-env:UserInstallation={profile_url}",
                "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
                abs_path,
            ],
            capture_output=True, text=True, timeout=180,
        )
        print(result.stdout)
        if result.returncode != 0:
            print(result.stderr, file=sys.stderr)
            raise RuntimeError("LibreOffice recalculation failed")

    # verify: every formula cell must have a real cached value, and none of
    # them may be an Excel error
    wb_formulas = openpyxl.load_workbook(abs_path, data_only=False)
    wb_values = openpyxl.load_workbook(abs_path, data_only=True)
    error_markers = ("#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
    problems = []
    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        for row in ws_f.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                v = ws_v[cell.coordinate].value
                if v is None:
                    problems.append(f"{sheet_name}!{cell.coordinate} formula produced no cached value")
                elif isinstance(v, str) and v in error_markers:
                    problems.append(f"{sheet_name}!{cell.coordinate} = {v}")
    if problems:
        raise RuntimeError("Recalculation problems found:\n" + "\n".join(problems[:50]))
    print("Recalculation clean: every formula has a real cached value, no errors.")
