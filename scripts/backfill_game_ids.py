"""
Backfills column Y ("Playgroup Game ID") onto already-logged rows in the
current season's Game Log tab, by matching each logged game against the
live, current-active-league /playgroup-games response -- the same feed and
the same matching algorithm (see computeLoggedMatches in app.js) already
used to decide what's missing from the Game Log. That scoping is
deliberate, not incidental: a historical game from an earlier season, or
any non-league game mixed into the current tab, simply has no candidate in
that response and is left blank -- this script only ever writes an ID for
a league game within the current active league, on purpose.

Once a row has an ID here, app.js's matcher can look games up by exact ID
instead of only inferring "same game" from player names + date + commander
-- the heuristic that caused the Game 13 bug fixed earlier this session.
Idempotent: a row that already has a value in column Y is never touched,
so re-running this after new games get logged (which write their own ID
automatically via add_game.py) only fills in what's still blank.

Usage:
    python scripts/backfill_game_ids.py             # dry run, report only
    python scripts/backfill_game_ids.py --apply      # write column Y
"""
import argparse
import json
import urllib.request
from datetime import datetime

import openpyxl

from xlsx_recalc import recalc
from season_config import CURRENT_SEASON_SHEET, XLSX_PATH

PLAYGROUP_GAMES_URL = "https://mtg-pod-validator-relay.mattdomi18.workers.dev/playgroup-games"
HEADER_ROW = 2


def normalize_commander(name):
    if not name:
        return ""
    return name.lower().split(",")[0].split("/")[0].strip()


def fetch_playgroup_games():
    req = urllib.request.Request(PLAYGROUP_GAMES_URL, headers={"User-Agent": "mtg-pod-validator-backfill-script"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.load(resp)


def find_col_map(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str):
            headers[v.strip()] = c
    return headers


def read_game_log_rows(ws, col_map):
    """[{row, gameNum, date, player, commander}, ...] for every data row."""
    rows = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        player = ws.cell(row=r, column=col_map["Player Name"]).value
        if not player:
            continue
        rows.append({
            "row": r,
            "gameNum": ws.cell(row=r, column=col_map["Game #"]).value,
            "date": ws.cell(row=r, column=col_map["Game Date"]).value,
            "player": player,
            "commander": ws.cell(row=r, column=col_map["Commander"]).value,
            "existingId": ws.cell(row=r, column=col_map["Playgroup Game ID"]).value if "Playgroup Game ID" in col_map else None,
        })
    return rows


def compute_matches(pg_games, log_rows):
    """Exact port of app.js's computeLoggedMatches: greedy, chronological,
    preferring a candidate whose logged commanders match the game's
    commanders, falling back to date proximity. Returns {playgroup_game_id:
    gameNum} for games that found a match. A logged game whose rows already
    carry an ID is excluded from the candidate pool entirely -- it's
    already claimed, and re-matching it here could only ever assign the
    same ID again at best or steal it from its rightful game at worst."""
    by_game_num = {}
    for row in log_rows:
        by_game_num.setdefault(row["gameNum"], []).append(row)

    logged_games = []
    for game_num, rows in by_game_num.items():
        if any(r["existingId"] for r in rows):
            continue
        logged_games.append({
            "gameNum": game_num,
            "players": set(r["player"] for r in rows),
            "commanders": {r["player"]: normalize_commander(r["commander"]) for r in rows},
            "date": rows[0]["date"] if isinstance(rows[0]["date"], datetime) else None,
            "claimed": False,
        })

    matches = {}
    for pg_game in sorted(pg_games, key=lambda g: g["date"]):
        pg_players = set(p["player"] for p in pg_game["participants"])
        pg_date = datetime.strptime(pg_game["date"], "%Y-%m-%d")

        best = None  # (logged, diff_days, commander_match)
        for logged in logged_games:
            if logged["claimed"]:
                continue
            if not pg_players.issubset(logged["players"]):
                continue

            diff_days = float("inf")
            if logged["date"]:
                diff_days = abs((logged["date"] - pg_date).total_seconds()) / 86400
                if diff_days > 1.5:
                    continue

            commander_match = all(
                logged["commanders"].get(p["player"]) == normalize_commander(p["commander"])
                for p in pg_game["participants"]
            )
            if (
                best is None
                or (commander_match and not best[2])
                or (commander_match == best[2] and diff_days < best[1])
            ):
                best = (logged, diff_days, commander_match)

        if best:
            best[0]["claimed"] = True
            matches[pg_game["playgroup_game_id"]] = best[0]["gameNum"]

    return matches


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="write column Y (default is dry-run report only)")
    args = parser.parse_args()

    print(f"Fetching live active-league games from {PLAYGROUP_GAMES_URL} ...")
    pg_data = fetch_playgroup_games()
    pg_games = pg_data["games"]
    print(f"{len(pg_games)} live game(s) in {pg_data.get('league', 'the active league')}.\n")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[CURRENT_SEASON_SHEET]
    col_map = find_col_map(ws)
    if "Playgroup Game ID" not in col_map:
        raise RuntimeError('Column "Playgroup Game ID" not found -- run add_game_id_column.py first.')

    log_rows = read_game_log_rows(ws, col_map)
    already_id_d = {r["gameNum"] for r in log_rows if r["existingId"]}
    matches = compute_matches(pg_games, log_rows)

    print(f"Already had an ID: {len(already_id_d)} game(s).")
    print(f"Newly matched:      {len(matches)} game(s).")
    unmatched_pg = [g for g in pg_games if g["playgroup_game_id"] not in matches]
    print(f"Still unmatched:    {len(unmatched_pg)} live game(s) -- these are what Games to Update shows as missing.\n")

    if matches:
        print("--- NEWLY MATCHED (would write) ---")
        for pg_id, game_num in sorted(matches.items(), key=lambda kv: kv[1]):
            print(f"  playgroup game {pg_id} -> Game #{game_num}")
        print()

    if not args.apply:
        print("Dry run only -- no changes written. Re-run with --apply to write column Y.")
        return

    # ---- apply: write column Y for every row of each newly matched game ----
    wb_write = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws_write = wb_write[CURRENT_SEASON_SHEET]
    col_map_write = find_col_map(ws_write)
    y_col = col_map_write["Playgroup Game ID"]
    game_col = col_map_write["Game #"]

    game_num_to_pg_id = {game_num: pg_id for pg_id, game_num in matches.items()}
    written = 0
    for r in range(HEADER_ROW + 1, ws_write.max_row + 1):
        game_num = ws_write.cell(row=r, column=game_col).value
        if game_num in game_num_to_pg_id:
            ws_write.cell(row=r, column=y_col).value = game_num_to_pg_id[game_num]
            written += 1

    wb_write.save(XLSX_PATH)
    print(f"Wrote {written} row(s) across {len(matches)} game(s) into column Y.")

    recalc(XLSX_PATH)


if __name__ == "__main__":
    main()
