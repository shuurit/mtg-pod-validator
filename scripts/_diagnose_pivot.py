import os
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "deck-strength.xlsx")

MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      Dim oSheets As Object
      Dim oSheet As Object
      Dim oPivots As Object
      Dim i As Integer, j As Integer

      oSheets = ThisComponent.getSheets()
      For i = 0 To oSheets.getCount() - 1
        oSheet = oSheets.getByIndex(i)
        oPivots = oSheet.getDataPilotTables()
        For j = 0 To oPivots.getCount() - 1
          oPivots.getByIndex(j).refresh()
        Next j
      Next i

      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""

tmp_copy = "/tmp/diag-deck-strength.xlsx"
subprocess.run(["cp", XLSX_PATH, tmp_copy], check=True)
abs_path = os.path.abspath(tmp_copy)

with tempfile.TemporaryDirectory(prefix="lo_profile_") as profile_dir:
    profile_url = Path(profile_dir).as_uri()
    subprocess.run(
        ["soffice", "--headless", "--terminate_after_init", f"-env:UserInstallation={profile_url}"],
        capture_output=True, text=True, timeout=60,
    )
    macro_dir = Path(profile_dir) / "user" / "basic" / "Standard"
    (macro_dir / "Module1.xba").write_text(MACRO)
    result = subprocess.run(
        [
            "soffice", "--headless", "--norestore",
            f"-env:UserInstallation={profile_url}",
            "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
            abs_path,
        ],
        capture_output=True, text=True, timeout=180,
    )
    print("STDOUT", result.stdout)
    print("STDERR", result.stderr, file=sys.stderr)

wb_f = openpyxl.load_workbook(abs_path, data_only=False)
wb_v = openpyxl.load_workbook(abs_path, data_only=True)
ws_f = wb_f["Deck Win Rates"]
ws_v = wb_v["Deck Win Rates"]
print("dims:", ws_f.dimensions)
for r in range(1, min(ws_f.max_row, 60) + 1):
    row_f = [ws_f.cell(row=r, column=c).value for c in range(1, 5)]
    row_v = [ws_v.cell(row=r, column=c).value for c in range(1, 5)]
    print(r, "F:", row_f, " V:", row_v)
