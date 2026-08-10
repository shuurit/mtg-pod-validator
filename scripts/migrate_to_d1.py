"""
One-time migration: reads deck-strength.xlsx and generates a SQL file that
populates the D1 database (see cloudflare-worker/schema.sql) with the same
data. Copies already-cached, already-correct values -- it does not
re-derive any formula, since those are already trusted (verified exact
against the spreadsheet's own cached output earlier this session).

Scope, deliberately: only Season 3 (CURRENT_SEASON_SHEET) games get
migrated into `games`/`game_results`. Player Adjusted Ranks' own formulas
hardcode 'Game Log Season 3' as their only source -- the live app has
never computed anything from Season 1/2's Game Logs, so this migration
doesn't invent a use for them either. All three season labels still get a
`seasons` row each (harmless, and needed as an FK target), but only
Season 3 gets real game data. Current Deck Strength's column C ("*Bracket
3", "*Bracket 4 review (...)") is free-text notes, not a clean field --
confirmed by direct inspection -- so `decks.bracket` is deliberately left
NULL by this migration; nothing in the current app reads it either.

Usage:
    python scripts/migrate_to_d1.py > migration_data.sql
    wrangler d1 execute mtg-pod-validator-db --remote --file=migration_data.sql
"""
import sys

import openpyxl

from season_config import XLSX_PATH, CURRENT_SEASON_SHEET

# Mirrors relay.js's USERNAME_TO_PLAYER (reversed) -- same duplicated-
# single-source-constant precedent already established by
# CURRENT_SEASON_SHEET between app.js and season_config.py.
PLAYER_TO_USERNAME = {
    "Becca": "Rebecca Dominguez",
    "Manny": "Thoros",
    "Mateo": "shuurit",
    "Ryan": "Ecthelion",
    "Michelle": "MLMyBelle",
    "Red": "Red",
}

CDS_SHEET = "Current Deck Strength"
SEASON_SHEETS = ["Game Log Season 1", "Game Log Season 2", "Game Log Season 3"]


def sql_str(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    return "NULL" if v is None else repr(float(v))


def sql_int(v):
    return "NULL" if v is None else str(int(v))


def read_players_and_decks(wb_values, wb_formulas):
    ws_v = wb_values[CDS_SHEET]
    ws_f = wb_formulas[CDS_SHEET]
    players = []  # [{name, decks: [...]}]
    current = None
    for r in range(2, ws_v.max_row + 1):
        name = ws_v.cell(row=r, column=1).value
        if not name:
            continue
        power_formula = ws_f.cell(row=r, column=2).value
        is_header = isinstance(power_formula, str) and power_formula.strip().startswith("=AVERAGE")
        if is_header:
            current = {"name": name, "decks": []}
            players.append(current)
            continue
        baseline = ws_v.cell(row=r, column=4).value
        playgroup_id = ws_v.cell(row=r, column=5).value
        current["decks"].append({
            "name": name,
            "baseline_power": baseline,
            "playgroup_deck_id": str(playgroup_id).strip() if playgroup_id not in (None, "") else None,
        })
    return players


def read_season3_games(wb_values):
    ws = wb_values[CURRENT_SEASON_SHEET]
    header_row = 2
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str):
            headers[v.strip()] = c

    def get(r, name):
        c = headers.get(name)
        return ws.cell(row=r, column=c).value if c else None

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        player = get(r, "Player Name")
        if not player:
            continue
        rows.append({
            "game_num": get(r, "Game #"),
            "date": get(r, "Game Date"),
            "player": player,
            "commander": get(r, "Commander"),
            "commander_strength": get(r, "Commander Strength"),
            "result": get(r, "Game Result"),
            "place": get(r, "Place"),
            "pod_size": get(r, "Pod Size"),
            "knockouts": get(r, "Knockouts"),
            "tov": get(r, "TOV"),
            "pop_off": get(r, "Pop-Off"),
            "disruptions": get(r, "Disruptions"),
            "recoveries": get(r, "Successful Recoveries"),
            "games_clearly_behind": get(r, "Games Clearly Behind"),
            "bracket": get(r, "Current Deck Bracket"),
            "playgroup_game_id": get(r, "Playgroup Game ID"),
            "J": get(r, "Adjusted Pod Size Win/Loss Score"),
            "K": get(r, "Knockout Score"),
            "L": get(r, "Deck Strength Comparison Differential"),
            "M": get(r, "Win Probability based on Deck Strength"),
            "N": get(r, "Player Score"),
            "O": get(r, "Normalized Player Score"),
            "Q": get(r, "Normalized TOV"),
            "U": get(r, "Deck Resilience Score"),
            "X": get(r, "Game Calculated Deck Strength"),
        })
    return rows


def normalize_commander(s):
    return (s or "").strip().lower()


def main():
    wb_values = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    wb_formulas = openpyxl.load_workbook(XLSX_PATH, data_only=False)

    players = read_players_and_decks(wb_values, wb_formulas)
    season_rows = read_season3_games(wb_values)

    out = []
    out.append("-- Generated by scripts/migrate_to_d1.py -- do not hand-edit, regenerate instead.")
    # No explicit BEGIN/COMMIT -- D1 rejects raw transaction statements (it
    # already batches an executed file atomically on its own; confirmed
    # the hard way, the whole file was rejected until these were removed).

    for label in ["Season 1", "Season 2", "Season 3"]:
        out.append(f"INSERT INTO seasons (id, label) VALUES ({label[-1]}, {sql_str(label)});")

    # deck_id lookup keyed by (player name, normalized deck name) so
    # game_results rows below can resolve their deck_id.
    deck_id_by_key = {}
    player_id_by_name = {}
    next_deck_id = 1
    for i, p in enumerate(players, start=1):
        player_id_by_name[p["name"]] = i
        username = PLAYER_TO_USERNAME.get(p["name"])
        out.append(
            f"INSERT INTO players (id, name, playgroup_username) VALUES "
            f"({i}, {sql_str(p['name'])}, {sql_str(username)});"
        )
        for d in p["decks"]:
            out.append(
                f"INSERT INTO decks (id, player_id, name, baseline_power, playgroup_deck_id, bracket, archived) "
                f"VALUES ({next_deck_id}, {i}, {sql_str(d['name'])}, {sql_num(d['baseline_power'])}, "
                f"{sql_str(d['playgroup_deck_id'])}, NULL, 0);"
            )
            deck_id_by_key[(p["name"], normalize_commander(d["name"]))] = next_deck_id
            next_deck_id += 1

    # games: one row per distinct Game #, deduped from the per-player rows.
    game_id_by_num = {}
    next_game_id = 1
    for row in season_rows:
        gn = row["game_num"]
        if gn in game_id_by_num:
            continue
        game_id_by_num[gn] = next_game_id
        played_at = row["date"].strftime("%Y-%m-%d") if row["date"] else None
        out.append(
            f"INSERT INTO games (id, season_id, game_num, played_at, pod_size, playgroup_game_id) "
            f"VALUES ({next_game_id}, 3, {sql_int(gn)}, {sql_str(played_at)}, {sql_int(row['pod_size'])}, "
            f"{sql_str(row['playgroup_game_id'])});"
        )
        next_game_id += 1

    unmatched = []
    for row in season_rows:
        game_id = game_id_by_num[row["game_num"]]
        player_id = player_id_by_name.get(row["player"])
        deck_id = deck_id_by_key.get((row["player"], normalize_commander(row["commander"])))
        if player_id is None or deck_id is None:
            unmatched.append(row)
            continue
        out.append(
            "INSERT INTO game_results (game_id, player_id, deck_id, commander_strength, result, place, "
            "knockouts, tov, pop_off, disruptions, recoveries, games_clearly_behind, bracket, "
            "j, k, l, m, n, o, q, u, x) VALUES ("
            f"{game_id}, {player_id}, {deck_id}, {sql_num(row['commander_strength'])}, {sql_int(row['result'])}, "
            f"{sql_int(row['place'])}, {sql_int(row['knockouts'])}, {sql_int(row['tov'])}, "
            f"{sql_int(row['pop_off'])}, {sql_int(row['disruptions'])}, {sql_int(row['recoveries'])}, "
            f"{sql_int(row['games_clearly_behind'])}, {sql_int(row['bracket'])}, "
            f"{sql_num(row['J'])}, {sql_num(row['K'])}, {sql_num(row['L'])}, {sql_num(row['M'])}, "
            f"{sql_num(row['N'])}, {sql_num(row['O'])}, {sql_num(row['Q'])}, {sql_num(row['U'])}, "
            f"{sql_num(row['X'])});"
        )

    print("\n".join(out))

    if unmatched:
        print(
            f"-- WARNING: {len(unmatched)} Game Log row(s) could not be matched to a "
            "player+deck and were skipped -- see stderr for details.",
            file=sys.stderr,
        )
        for row in unmatched:
            print(f"  unmatched: game {row['game_num']} player={row['player']!r} commander={row['commander']!r}", file=sys.stderr)


if __name__ == "__main__":
    main()
