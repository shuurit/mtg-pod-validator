"""
Adds a game to 'Game Log Season 3' in deck-strength.xlsx from a JSON payload,
then forces a full LibreOffice recalculation so every formula in the
workbook (including Player Adjusted Ranks) reflects the new game.

Expects the payload as JSON in the GAME_PAYLOAD environment variable:
{
  "date": "2026-08-02",
  "podSize": 3,
  "participants": [
    {
      "player": "Manny", "commander": "Kruphix, God of Horizons",
      "strength": 2.3, "result": "win", "place": 1, "knockouts": 2,
      "tov": 8, "popOff": 1, "disruptions": 0, "recoveries": 0,
      "gamesClearlyBehind": 0, "bracket": 2
    },
    ...
  ]
}

Participants should already be sorted winner-first / place-ascending
(matches the sheet's existing convention) -- this script does not resort.
"""
import copy
import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "deck-strength.xlsx")
SHEET_NAME = "Game Log Season 3"
HEADER_ROW = 2
COLUMNS = "ABCDEFGHIJKLMNOPQRSTUVWX"


def find_col_map(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(v, str):
            headers[v.strip()] = c
    return headers


def find_next_slot(ws, col_map):
    game_col = col_map["Game #"]
    last_row = HEADER_ROW
    max_game_num = 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=game_col).value
        if v is None:
            continue
        last_row = r
        if isinstance(v, (int, float)) and v > max_game_num:
            max_game_num = int(v)
    return last_row + 1, max_game_num + 1


def main():
    payload = json.loads(os.environ["GAME_PAYLOAD"])
    date_str = payload["date"]
    pod_size = payload["podSize"]
    participants = payload["participants"]
    if len(participants) != pod_size:
        raise ValueError(f"podSize={pod_size} but got {len(participants)} participants")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    ws = wb[SHEET_NAME]
    col_map = find_col_map(ws)
    start_row, game_num = find_next_slot(ws, col_map)

    game_date = datetime.strptime(date_str, "%Y-%m-%d")
    rows = list(range(start_row, start_row + pod_size))

    # style reference: the most recently populated row, so new rows visually
    # match whatever the sheet already looks like (formats, fonts, etc.)
    style_ref_row = start_row - 1

    for i, (row, p) in enumerate(zip(rows, participants)):
        other_rows = [r for r in rows if r != row]
        result = 1 if p["result"] == "win" else 0
        values = {
            "A": game_date,
            "B": game_num,
            "C": p["player"],
            "D": p["commander"],
            "E": p["strength"],
            "F": result,
            "G": p["place"],
            "H": pod_size,
            "I": p["knockouts"],
            "J": f"=F{row}-(1/H{row})",
            "K": f"=((I{row}-((H{row}-1)/H{row}))-(-5/6))/((5-(5/6))-(-5/6))",
            "L": f"=E{row}-(({'+'.join(f'E{o}' for o in other_rows)})/{pod_size - 1})",
            "M": f"=0.5+(L{row}*0.09)",
            "N": f"=((H{row}-(G{row}-1))/H{row})*(IF(I{row}<>0,(I{row}+H{row})/H{row},1))",
            "O": f"=(N{row}-(1/6))/(10/6)",
            "P": p["tov"],
            "Q": f"=IF(F{row}=1,(((1-((P{row}-3)/15))*0.5)+0.5),((P{row}/18)*0.5))",
            "R": p["popOff"],
            "S": p["disruptions"],
            "T": p["recoveries"],
            "U": f"=IFERROR(T{row}/S{row}, 1)",
            "V": p["gamesClearlyBehind"],
            "W": p["bracket"],
            "X": f"=((O{row}*0.3)+(Q{row}*0.175)+(R{row}*0.175)+(U{row}*0.175)+((1-V{row})*0.175))+W{row}",
        }
        for col_letter, val in values.items():
            col = openpyxl.utils.column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=col)
            src_cell = ws.cell(row=style_ref_row, column=col)
            cell.value = val
            if src_cell.has_style:
                cell._style = copy.copy(src_cell._style)

    wb.save(XLSX_PATH)
    print(f"Wrote Game #{game_num} at rows {rows[0]}-{rows[-1]}.")

    # openpyxl's save strips cached formula values workbook-wide -- force a
    # full LibreOffice recalculation to restore them everywhere, not just
    # the rows we touched.
    recalc()


RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def recalc():
    # Plain `soffice --convert-to` does NOT reliably force a full formula
    # recalculation -- confirmed the hard way: it silently produced a file
    # where every formula cell across the whole workbook was simply missing
    # (not an error value, just absent), because openpyxl's save() had
    # stripped their cached values and convert-to never recomputed them.
    # A macro that explicitly calls calculateAll() before store() is the
    # reliable way to force it.
    abs_path = os.path.abspath(XLSX_PATH)

    with tempfile.TemporaryDirectory(prefix="lo_profile_") as profile_dir:
        profile_url = Path(profile_dir).as_uri()

        # bootstrap the profile so LibreOffice has created its basic/Standard dir
        boot = subprocess.run(
            ["soffice", "--headless", "--terminate_after_init", f"-env:UserInstallation={profile_url}"],
            capture_output=True, text=True, timeout=60,
        )
        if boot.returncode != 0:
            print(boot.stderr, file=sys.stderr)
            raise RuntimeError("LibreOffice failed to initialize its profile")

        macro_dir = Path(profile_dir) / "user" / "basic" / "Standard"
        if not macro_dir.exists():
            raise RuntimeError("LibreOffice did not create a usable profile directory")
        (macro_dir / "Module1.xba").write_text(RECALCULATE_MACRO)

        result = subprocess.run(
            [
                "soffice", "--headless", "--norestore",
                f"-env:UserInstallation={profile_url}",
                "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
                abs_path,
            ],
            capture_output=True, text=True, timeout=180,
        )
        print(result.stdout)
        if result.returncode != 0:
            print(result.stderr, file=sys.stderr)
            raise RuntimeError("LibreOffice recalculation failed")

    # verify: every formula cell must have a real cached value, and none of
    # them may be an Excel error
    wb_formulas = openpyxl.load_workbook(abs_path, data_only=False)
    wb_values = openpyxl.load_workbook(abs_path, data_only=True)
    error_markers = ("#DIV/0!", "#REF!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
    problems = []
    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]
        for row in ws_f.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                v = ws_v[cell.coordinate].value
                if v is None:
                    problems.append(f"{sheet_name}!{cell.coordinate} formula produced no cached value")
                elif isinstance(v, str) and v in error_markers:
                    problems.append(f"{sheet_name}!{cell.coordinate} = {v}")
    if problems:
        raise RuntimeError("Recalculation problems found:\n" + "\n".join(problems[:50]))
    print("Recalculation clean: every formula has a real cached value, no errors.")


if __name__ == "__main__":
    main()
