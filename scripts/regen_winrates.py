"""
Refreshes the player_adjusted_win_rate side of player-win-rates.json from
the (freshly recalculated) Player Adjusted Ranks tab. Leaves
playgroup_win_rate alone -- that side is tallied from playgroup.gg's active
league game history directly and doesn't change just because a game got
logged into the spreadsheet.
"""
import json
import os

import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), "..", "deck-strength.xlsx")
WINRATES_PATH = os.path.join(os.path.dirname(__file__), "..", "player-win-rates.json")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Player Adjusted Ranks"]

    adjusted = {}
    for row in range(2, ws.max_row + 1):
        player = ws.cell(row=row, column=1).value
        if not player:
            continue
        adjusted[player] = {
            "rate": ws.cell(row=row, column=2).value,
            "wins": ws.cell(row=row, column=3).value,
            "losses": ws.cell(row=row, column=4).value,
        }

    with open(WINRATES_PATH) as f:
        data = json.load(f)

    for entry in data["players"]:
        player = entry["player"]
        adj = adjusted.get(player)
        if not adj:
            continue
        games_logged = (adj["wins"] or 0) + (adj["losses"] or 0)
        entry["player_adjusted_record"] = f"{adj['wins']}-{adj['losses']}"
        if games_logged:
            entry["player_adjusted_win_rate"] = round(adj["rate"] * 100, 2)
            entry.pop("player_adjusted_note", None)
        else:
            entry["player_adjusted_win_rate"] = None
            entry["player_adjusted_note"] = "No games logged in Game Log Season 3"

    with open(WINRATES_PATH, "w") as f:
        json.dump(data, f, indent=2)
    print("player-win-rates.json refreshed from Player Adjusted Ranks.")


if __name__ == "__main__":
    main()
